#!/usr/bin/env python3
"""
run_pipeline_excel.py — Pipeline NUMT sans MongoDB, sortie Excel multi-feuilles.

Chaque étape intermédiaire est sauvegardée dans une feuille Excel distincte,
ce qui permet d'inspecter visuellement où les variants sont perdus ou modifiés.

OBJECTIF SCIENTIFIQUE
=====================
Identifier parmi les insertions ClinVar pathogènes (>50 bp) les variants qui :

  CIBLE PRINCIPALE :
    Match mitochondrie  +  Pas de match nucléaire  →  NOVEL_NUMT_CANDIDATE
    Le fragment mtDNA n'est PAS dans le génome de référence hg38.
    → Insertion de novo pathogène probable (cf. cas index TSC2 / Bhatt et al.)

  À DOCUMENTER :
    Match mitochondrie  +  Match nucléaire  →  CONFIRMED_NUMT / KNOWN_REF_NUMT
    La séquence existe déjà dans hg38 (NUMT de référence fixé dans la population)
    ou correspond au site ClinVar rapporté.

FEUILLES EXCEL
==============
  1_ClinVar_Raw       — insertions >50 bp filtrées depuis ClinVar (sortie étape 1)
  2_Sequences         — après parsing HGVS et extraction de séquence (étape 2)
  3_Deduplicated      — 1 ligne par AlleleID, préférence GRCh38 (remplace MongoDB)
  4_BLAST_Mito        — résultats BLAST vs génome mitochondrial (1 ligne par HSP)
  5_BLAST_Genome      — résultats BLAST vs génome humain pour les hits mito
  6_Classification    — classification finale par variant (numt_class)
  NOVEL_CANDIDATES    — vue filtrée : hit mito, PAS de hit nucléaire  ← PRIORITÉ
  MITO_AND_NUCLEAR    — vue filtrée : hit mito ET hit nucléaire
  7_Mechanism         — mécanisme d'insertion prédit (TPRT/NAHR/NHEJ/MMEJ/MMBIR)

USAGE
=====
  # Pipeline complet (BLAST génome en mode remote NCBI) :
  python run_pipeline_excel.py --email your@email.com

  # Sauter le BLAST génome (classification basée uniquement sur mito) :
  python run_pipeline_excel.py --skip-genome

  # BLAST génome local (plus rapide, nécessite base hg38 locale) :
  python run_pipeline_excel.py --genome-db /path/to/blast_db/hg38

  # Reprendre depuis l'étape 3 (étapes 1-2 déjà faites) :
  python run_pipeline_excel.py --from-step 3 --email your@email.com

  # Forcer le re-téléchargement ClinVar :
  python run_pipeline_excel.py --force
"""

# ─── Standard-library ────────────────────────────────────────────────────────
import argparse
import importlib.util
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

# ─── Third-party ─────────────────────────────────────────────────────────────
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from Bio import Entrez, SeqIO
    BIOPYTHON_AVAILABLE = True
except ImportError:
    BIOPYTHON_AVAILABLE = False

try:
    from Bio.Blast import NCBIWWW, NCBIXML
    BIOPYTHON_BLAST_AVAILABLE = True
except ImportError:
    BIOPYTHON_BLAST_AVAILABLE = False


# ─── Chemins projet ──────────────────────────────────────────────────────────

SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
DATA_RAW     = PROJECT_ROOT / "02_data" / "raw"
DATA_PROC    = PROJECT_ROOT / "02_data" / "processed"

# Entrées/sorties des étapes existantes
CLINVAR_RAW_TSV   = DATA_PROC / "clinvar_insertions_gt50bp.tsv"
SEQUENCES_TSV     = DATA_PROC / "clinvar_insertions_with_sequences.tsv"

# Fichiers BLAST
MITO_FASTA        = DATA_RAW  / "NC_012920.1.fasta"
BLAST_DB_DIR      = DATA_RAW  / "blast_db" / "mito"
BLAST_DB_NAME     = "mito_db"
QUERY_FASTA       = DATA_PROC / "excel_pipeline_query.fasta"
BLAST_MITO_TSV    = DATA_PROC / "excel_blast_mito.tsv"
BLAST_GENOME_TSV  = DATA_PROC / "excel_blast_genome.tsv"
GENOME_CACHE_DIR  = DATA_PROC / "blast_genome_cache"

# Sortie finale
OUTPUT_EXCEL      = DATA_PROC / "numt_pipeline_results.xlsx"

# Sorties mécanisme
MECHANISM_JSON    = DATA_PROC / "mechanism_analysis.json"
MECHANISM_TSV     = DATA_PROC / "mechanism_summary.tsv"

# Sorties LAST validation
LAST_MITO_TSV     = DATA_PROC / "excel_last_mito.tsv"
LAST_COMPARE_TSV  = DATA_PROC / "blast_vs_last_comparison.tsv"
LAST_REPORT_TXT   = DATA_PROC / "blast_vs_last_report.txt"
LAST_WORKDIR      = DATA_PROC / "last_work"


# ─── Colonnes BLAST tabular (-outfmt 6) ──────────────────────────────────────

BLAST_COLS = [
    "qseqid", "sseqid", "pident", "length", "mismatch", "gapopen",
    "qstart",  "qend",  "sstart",  "send",  "evalue",  "bitscore",
    "qlen",    "slen",
]

# ─── Annotation génome mitochondrial (rCRS NC_012920.1) ──────────────────────

MTDNA_FEATURES = [
    ("D-loop",         16024, 16569, "control_region"),
    ("D-loop",             1,   576, "control_region"),
    ("12S rRNA",         648,  1601, "rRNA"),
    ("16S rRNA",        1671,  3229, "rRNA"),
    ("MT-ND1",          3307,  4262, "protein_coding"),
    ("MT-ND2",          4470,  5511, "protein_coding"),
    ("MT-CO1",          5904,  7445, "protein_coding"),
    ("MT-CO2",          7586,  8269, "protein_coding"),
    ("MT-ATP8",         8366,  8572, "protein_coding"),
    ("MT-ATP6",         8527,  9207, "protein_coding"),
    ("MT-CO3",          9207,  9990, "protein_coding"),
    ("MT-ND3",         10059, 10404, "protein_coding"),
    ("MT-ND4L",        10470, 10766, "protein_coding"),
    ("MT-ND4",         10760, 12137, "protein_coding"),
    ("MT-ND5",         12337, 14148, "protein_coding"),
    ("MT-ND6",         14149, 14673, "protein_coding"),
    ("MT-CYB",         14747, 15887, "protein_coding"),
]

# Gènes de maladies connues — marqués dans les résultats
DISEASE_GENES = {
    "TSC1", "TSC2", "NF1", "NF2", "BRCA1", "BRCA2", "APC", "TP53",
    "RB1", "VHL", "MLH1", "MSH2", "MSH6", "PMS2", "CFTR", "DMD",
    "FBN1", "PKD1", "PKD2", "ATM", "PTEN",
}


# ─── Logging ─────────────────────────────────────────────────────────────────

def log(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}")


# ─── ÉTAPE 1 & 2 : appel des scripts existants ──────────────────────────────

def run_step(script: str, extra_args: list[str] = None) -> bool:
    """Lance un script existant de la pipeline via subprocess."""
    script_path = SCRIPT_DIR / script
    cmd = [sys.executable, str(script_path)] + (extra_args or [])
    log(f"→ {' '.join(cmd)}")
    result = subprocess.run(cmd, check=False)
    return result.returncode == 0


# ─── ÉTAPE 3 : dédoublonnage AlleleID ────────────────────────────────────────

def dedup_by_allele(df: pd.DataFrame) -> pd.DataFrame:
    """
    Une ligne par AlleleID, préférence GRCh38 sur GRCh37.

    ClinVar contient une ligne par (variant, assembly), donc le même AlleleID
    peut apparaître deux fois. On garde GRCh38 en priorité.
    """
    allele_col = "#AlleleID" if "#AlleleID" in df.columns else "AlleleID"
    if allele_col not in df.columns:
        log("WARNING — Colonne AlleleID introuvable, pas de dédoublonnage.")
        return df

    # Normalisation du nom de colonne
    df = df.rename(columns={"#AlleleID": "AlleleID"})

    n_before = len(df)
    if "Assembly" in df.columns:
        df = df.sort_values("Assembly", ascending=False, na_position="last")
    df = df.drop_duplicates(subset=["AlleleID"], keep="first")
    log(f"Dédoublonnage : {n_before:,} lignes → {len(df):,} variants uniques")
    return df.reset_index(drop=True)


# ─── ÉTAPE 4a : génération du FASTA de requête ───────────────────────────────

def build_query_fasta(df: pd.DataFrame, rcrs_seq: str | None = None) -> int:
    """
    Génère le FASTA de requête à partir du DataFrame dédoublonné.

    Inclut :
      - DIRECT_SEQUENCE  : séquence extraite directement du nom HGVS
      - DIRECT_MITO_REF  : séquence récupérée depuis la rCRS via mito_ref_range

    Format des headers : >AlleleID|GeneSymbol|ClinSig|LengthBp
    """
    QUERY_FASTA.parent.mkdir(parents=True, exist_ok=True)

    n = 0
    with open(QUERY_FASTA, "w") as fh:
        for _, row in df.iterrows():
            status = str(row.get("extraction_status", ""))
            seq    = row.get("extracted_sequence", None)

            # Cas 1 : séquence directement disponible
            if status == "DIRECT_SEQUENCE" and isinstance(seq, str) and seq.strip():
                pass  # seq déjà définie

            # Cas 2 : référence mitochondriale → extraire depuis rCRS
            elif status == "DIRECT_MITO_REF" and rcrs_seq:
                mito_range = str(row.get("mito_ref_range", ""))
                if "-" in mito_range:
                    try:
                        start, end = mito_range.split("-")
                        # Coordonnées 1-based → indices Python 0-based
                        seq = rcrs_seq[int(start) - 1 : int(end)]
                    except (ValueError, IndexError):
                        seq = None
                else:
                    seq = None
            else:
                continue

            if not seq:
                continue

            allele_id = row.get("AlleleID", "NA")
            gene      = str(row.get("GeneSymbol", "NA") or "NA").replace("|", "/")
            clinsig   = str(row.get("ClinicalSignificance", "NA") or "NA")
            clinsig   = clinsig.replace("|", "/").replace(" ", "_")
            length_bp = len(seq)

            header = f">{allele_id}|{gene}|{clinsig}|{length_bp}"
            fh.write(f"{header}\n")
            for i in range(0, len(seq), 80):
                fh.write(seq[i:i + 80] + "\n")
            n += 1

    log(f"FASTA de requête : {n:,} séquences écrites → {QUERY_FASTA}")
    return n


# ─── ÉTAPE 4b : téléchargement rCRS et construction base BLAST ───────────────

def ensure_rcrs(email: str | None = None) -> str | None:
    """
    Télécharge la rCRS (NC_012920.1) si absente, retourne la séquence en str.
    """
    if MITO_FASTA.exists():
        log(f"rCRS déjà présente : {MITO_FASTA}")
    else:
        if not BIOPYTHON_AVAILABLE:
            log("FATAL — biopython requis pour télécharger la rCRS. pip install biopython")
            return None
        log("Téléchargement rCRS (NC_012920.1) depuis NCBI ...")
        Entrez.email = email or "numt_pipeline@example.com"
        try:
            handle = Entrez.efetch(
                db="nucleotide", id="NC_012920.1", rettype="fasta", retmode="text"
            )
            content = handle.read()
            handle.close()
            MITO_FASTA.parent.mkdir(parents=True, exist_ok=True)
            MITO_FASTA.write_text(content)
            log(f"rCRS sauvegardée ({len(content):,} octets)")
        except Exception as exc:
            log(f"FATAL — Échec téléchargement rCRS : {exc}")
            return None

    # Lecture de la séquence
    try:
        record = next(SeqIO.parse(str(MITO_FASTA), "fasta"))
        return str(record.seq)
    except Exception as exc:
        log(f"WARNING — Impossible de lire la rCRS : {exc}")
        return None


def ensure_blast_db() -> Path | None:
    """Construit la base BLAST mitochondriale si nécessaire."""
    if shutil.which("makeblastdb") is None:
        log("FATAL — makeblastdb introuvable. brew install blast")
        return None

    BLAST_DB_DIR.mkdir(parents=True, exist_ok=True)
    db_path = BLAST_DB_DIR / BLAST_DB_NAME

    if (BLAST_DB_DIR / f"{BLAST_DB_NAME}.nsq").exists() or \
       (BLAST_DB_DIR / f"{BLAST_DB_NAME}.ndb").exists():
        log(f"Base BLAST mito déjà présente : {db_path}")
        return db_path

    log("Construction de la base BLAST mitochondriale ...")
    cmd = [
        "makeblastdb", "-in", str(MITO_FASTA), "-dbtype", "nucl",
        "-out", str(db_path), "-title", "rCRS", "-parse_seqids",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        log(f"FATAL — makeblastdb a échoué : {result.stderr}")
        return None
    log("Base BLAST mito construite.")
    return db_path


# ─── Classification et annotation mtDNA (réutilisées par LAST validation) ────

def classify_hit(row):
    """Classe un hit mito par tier de confiance."""
    if row["pident"] >= 90 and row["length"] >= 50 and row["evalue"] <= 1e-5:
        return "STRONG_HIT"
    elif row["pident"] >= 80 and row["length"] >= 40 and row["evalue"] <= 1e-3:
        return "MODERATE_HIT"
    return "WEAK_HIT"


def annotate_region(sstart, send):
    """Retourne (region_name, region_type) pour une position sur la rCRS."""
    hit_start, hit_end = min(sstart, send), max(sstart, send)
    best_name, best_type, best_overlap = "intergenic", "intergenic", 0
    for name, fs, fe, ft in MTDNA_FEATURES:
        overlap = max(0, min(hit_end, fe) - max(hit_start, fs) + 1)
        if overlap > best_overlap:
            best_overlap, best_name, best_type = overlap, name, ft
    return best_name, best_type


# ─── ÉTAPE 4c : exécution BLAST vs mito ──────────────────────────────────────

def run_blast_mito(db_path: Path) -> pd.DataFrame:
    """Lance blastn vs rCRS et retourne les résultats annotés."""
    if shutil.which("blastn") is None:
        log("FATAL — blastn introuvable. brew install blast")
        return pd.DataFrame()

    # Auto-détection blastn-short vs blastn
    lengths = []
    with open(QUERY_FASTA) as fh:
        cur = 0
        for line in fh:
            if line.startswith(">"):
                if cur > 0:
                    lengths.append(cur)
                cur = 0
            else:
                cur += len(line.strip())
        if cur > 0:
            lengths.append(cur)

    if not lengths:
        log("FATAL — FASTA de requête vide.")
        return pd.DataFrame()

    median_len = sorted(lengths)[len(lengths) // 2]
    task = "blastn-short" if median_len < 100 else "blastn"
    log(f"BLAST mito — task={task}, médiane des requêtes={median_len} bp")

    outfmt = "6 qseqid sseqid pident length mismatch gapopen qstart qend sstart send evalue bitscore qlen slen"
    cmd = [
        "blastn",
        "-query",          str(QUERY_FASTA),
        "-db",             str(db_path),
        "-out",            str(BLAST_MITO_TSV),
        "-task",           task,
        "-evalue",         "1e-3",
        "-word_size",      "7",
        "-perc_identity",  "80",
        "-qcov_hsp_perc",  "50",
        "-outfmt",         outfmt,
        "-num_threads",    "4",
    ]
    log("Lancement BLAST vs mitochondrie ...")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        log(f"FATAL — blastn a échoué : {result.stderr}")
        return pd.DataFrame()

    if not BLAST_MITO_TSV.exists() or BLAST_MITO_TSV.stat().st_size == 0:
        log("BLAST mito : aucun hit trouvé.")
        return pd.DataFrame(columns=BLAST_COLS + [
            "allele_id", "gene_symbol", "hit_tier",
            "mtdna_region", "mtdna_region_type", "query_coverage", "combined_score",
        ])

    df = pd.read_csv(BLAST_MITO_TSV, sep="\t", header=None, names=BLAST_COLS)
    log(f"BLAST mito : {len(df):,} HSPs trouvés")

    # Extraction AlleleID et gene depuis le header FASTA
    df["allele_id"] = df["qseqid"].apply(
        lambda x: int(x.split("|")[0]) if "|" in str(x) else None
    )
    df["gene_symbol"] = df["qseqid"].apply(
        lambda x: x.split("|")[1] if "|" in str(x) and len(x.split("|")) > 1 else None
    )

    df["hit_tier"] = df.apply(classify_hit, axis=1)

    regions = df.apply(lambda r: annotate_region(int(r["sstart"]), int(r["send"])), axis=1)
    df["mtdna_region"]      = regions.apply(lambda x: x[0])
    df["mtdna_region_type"] = regions.apply(lambda x: x[1])

    df["query_coverage"] = (df["length"] / df["qlen"]) * 100.0
    df["combined_score"] = df["pident"] * (df["length"] / df["qlen"]) / 100.0

    return df


# ─── ÉTAPE 5 : BLAST vs génome humain ────────────────────────────────────────

def build_genome_fasta(df_mito: pd.DataFrame, df_dedup: pd.DataFrame) -> Path:
    """
    Génère un FASTA pour le BLAST génome, uniquement avec les variants
    qui ont un hit mitochondrial.
    """
    mito_allele_ids = set(df_mito["allele_id"].dropna().astype(int))
    df_hits = df_dedup[
        df_dedup["AlleleID"].astype(str).apply(
            lambda x: int(x) if x.isdigit() else -1
        ).isin(mito_allele_ids)
    ]

    genome_fasta = DATA_PROC / "excel_pipeline_genome_query.fasta"
    n = 0
    with open(genome_fasta, "w") as fh:
        for _, row in df_hits.iterrows():
            seq = row.get("extracted_sequence", None)
            if not isinstance(seq, str) or not seq.strip():
                continue
            allele_id = row.get("AlleleID", "NA")
            gene      = str(row.get("GeneSymbol", "NA") or "NA")
            chrom     = str(row.get("Chromosome", "NA") or "NA")
            start_pos = str(row.get("Start", "NA") or "NA")
            fh.write(f">{allele_id}|{gene}|{chrom}|{start_pos}\n")
            for i in range(0, len(seq), 80):
                fh.write(seq[i:i + 80] + "\n")
            n += 1

    log(f"FASTA pour BLAST génome : {n:,} séquences (variants avec hit mito)")
    return genome_fasta


def run_blast_genome_local(genome_fasta: Path, genome_db: Path) -> pd.DataFrame:
    """BLAST vs hg38 en mode local."""
    if shutil.which("blastn") is None:
        log("FATAL — blastn introuvable.")
        return pd.DataFrame()

    outfmt = "6 qseqid sseqid pident length mismatch gapopen qstart qend sstart send evalue bitscore qlen slen"
    cmd = [
        "blastn",
        "-query",          str(genome_fasta),
        "-db",             str(genome_db),
        "-out",            str(BLAST_GENOME_TSV),
        "-task",           "blastn",
        "-evalue",         "1e-5",
        "-perc_identity",  "80",
        "-qcov_hsp_perc",  "50",
        "-outfmt",         outfmt,
        "-num_threads",    "4",
    ]
    log("BLAST vs génome humain (local) ...")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        log(f"FATAL — blastn génome a échoué : {result.stderr}")
        return pd.DataFrame()
    return _parse_genome_blast()


def run_blast_genome_remote(genome_fasta: Path, email: str) -> pd.DataFrame:
    """BLAST vs hg38 en mode remote (NCBI qblast), avec cache par AlleleID."""
    if not BIOPYTHON_BLAST_AVAILABLE:
        log("FATAL — biopython[blast] requis. pip install biopython")
        return pd.DataFrame()

    GENOME_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    Entrez.email = email

    # Lecture des séquences depuis le FASTA
    sequences = {}
    with open(genome_fasta) as fh:
        current_id = None
        current_seq = []
        for line in fh:
            if line.startswith(">"):
                if current_id:
                    sequences[current_id] = "".join(current_seq)
                current_id = line.strip()[1:].split("|")[0]
                current_seq = []
            else:
                current_seq.append(line.strip())
        if current_id:
            sequences[current_id] = "".join(current_seq)

    log(f"BLAST remote vs hg38 : {len(sequences)} séquences à soumettre")
    log("NCBI rate limit : 1 requête / 10 secondes — patience ...")

    all_rows = []
    for i, (allele_id, seq) in enumerate(sequences.items()):
        cache_file = GENOME_CACHE_DIR / f"{allele_id}.xml"

        if cache_file.exists():
            log(f"  [{i+1}/{len(sequences)}] AlleleID {allele_id} : résultat en cache")
            xml_data = cache_file.read_text()
        else:
            log(f"  [{i+1}/{len(sequences)}] AlleleID {allele_id} : soumission NCBI ...")
            try:
                handle = NCBIWWW.qblast(
                    "blastn", "nt", seq,
                    entrez_query="Homo sapiens[ORGN]",
                    expect=1e-5,
                    perc_ident=80,
                )
                xml_data = handle.read()
                handle.close()
                cache_file.write_text(xml_data)
                time.sleep(3)  # respect NCBI rate limit (minimum 3s between submissions)
            except Exception as exc:
                log(f"  WARNING — qblast a échoué pour AlleleID {allele_id} : {exc}")
                continue

        # Parsing XML
        try:
            from io import StringIO as SIO
            blast_records = list(NCBIXML.parse(SIO(xml_data)))
            for rec in blast_records:
                for aln in rec.alignments:
                    for hsp in aln.hsps:
                        # Extraire chromosome depuis le titre
                        title = aln.title.lower()
                        is_nuclear = not any(
                            x in title for x in ["mitochondri", "chrom m", "chrm", "nc_012920"]
                        )
                        all_rows.append({
                            "allele_id":   allele_id,
                            "subject":     aln.hit_id,
                            "subject_title": aln.title[:80],
                            "pident":      hsp.identities / hsp.align_length * 100,
                            "length":      hsp.align_length,
                            "evalue":      hsp.expect,
                            "bitscore":    hsp.bits,
                            "qstart":      hsp.query_start,
                            "qend":        hsp.query_end,
                            "sstart":      hsp.sbjct_start,
                            "send":        hsp.sbjct_end,
                            "is_nuclear":  is_nuclear,
                        })
        except Exception as exc:
            log(f"  WARNING — parsing XML échoué pour {allele_id} : {exc}")

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    df.to_csv(BLAST_GENOME_TSV, sep="\t", index=False)
    log(f"BLAST génome remote : {len(df):,} HSPs")
    return df


def _parse_genome_blast() -> pd.DataFrame:
    """Parse le TSV de sortie du BLAST génome local."""
    if not BLAST_GENOME_TSV.exists() or BLAST_GENOME_TSV.stat().st_size == 0:
        log("BLAST génome : aucun hit.")
        return pd.DataFrame()

    df = pd.read_csv(BLAST_GENOME_TSV, sep="\t", header=None, names=BLAST_COLS)
    df["allele_id"] = df["qseqid"].apply(
        lambda x: x.split("|")[0] if "|" in str(x) else x
    )
    df["gene_symbol"] = df["qseqid"].apply(
        lambda x: x.split("|")[1] if "|" in str(x) and len(x.split("|")) > 1 else None
    )
    df["chromosome"] = df["qseqid"].apply(
        lambda x: x.split("|")[2] if "|" in str(x) and len(x.split("|")) > 2 else None
    )
    # Identifier les hits nucléaires (exclure chrM)
    df["is_nuclear"] = ~df["sseqid"].str.lower().str.contains(
        "nc_012920|mito|chrm", na=False
    )
    log(f"BLAST génome : {len(df):,} HSPs parsés")
    return df


# ─── ÉTAPE 6 : classification finale ─────────────────────────────────────────

def classify_variants(
    df_dedup: pd.DataFrame,
    df_mito: pd.DataFrame,
    df_genome: pd.DataFrame | None,
) -> pd.DataFrame:
    """
    Assigne une classe NUMT à chaque variant selon les résultats BLAST.

    numt_class :
      NOVEL_NUMT_CANDIDATE   — hit mito + PAS de hit nucléaire  ← CIBLE
      CONFIRMED_NUMT         — hit mito + hit nucléaire au locus ClinVar
      KNOWN_REFERENCE_NUMT   — hit mito + hit nucléaire ailleurs dans hg38
      MITO_HIT_ONLY          — hit mito, génome non testé (--skip-genome)
      NO_MITO_HIT            — aucun hit mitochondrial
      DIRECT_MITO_REF        — référence mito explicite dans le nom HGVS
    """
    df = df_dedup.copy()
    allele_col = "AlleleID"

    # Normalisation AlleleID en int
    df["allele_id_int"] = pd.to_numeric(df[allele_col], errors="coerce")

    # Variants avec hit mito
    if not df_mito.empty:
        mito_ids = set(df_mito["allele_id"].dropna().astype(str).astype(float).astype(int))
    else:
        mito_ids = set()

    # Meilleur hit mito par variant
    mito_best = {}
    if not df_mito.empty:
        for aid, grp in df_mito.groupby("allele_id"):
            if aid is None:
                continue
            best = grp.sort_values("combined_score", ascending=False).iloc[0]
            mito_best[int(aid)] = {
                "mito_pident":     round(float(best["pident"]), 2),
                "mito_evalue":     float(best["evalue"]),
                "mito_hit_tier":   best["hit_tier"],
                "mito_region":     best["mtdna_region"],
                "mito_region_type": best["mtdna_region_type"],
                "mito_combined_score": round(float(best["combined_score"]), 4),
                "mito_n_hsps":     len(grp),
            }

    # Variants avec hit nucléaire
    nuclear_ids = set()
    nuclear_at_locus = set()

    if df_genome is not None and not df_genome.empty and "is_nuclear" in df_genome.columns:
        # Handle both boolean and string "True"/"False" (TSV loaded with dtype=str)
        nuclear_hits = df_genome[df_genome["is_nuclear"].astype(str).str.lower() == "true"]
        if not nuclear_hits.empty:
            nuclear_ids = set(
                nuclear_hits["allele_id"].dropna().astype(str).apply(
                    lambda x: int(float(x)) if x.replace(".", "").isdigit() else None
                ).dropna().astype(int)
            )
            # Vérifier si le hit correspond au locus ClinVar rapporté
            for aid_str, grp in nuclear_hits.groupby("allele_id"):
                try:
                    aid = int(float(aid_str))
                except (ValueError, TypeError):
                    continue
                # Comparer le chromosome du hit avec celui de ClinVar
                clinvar_row = df[df["allele_id_int"] == aid]
                if not clinvar_row.empty:
                    clinvar_chrom = str(clinvar_row.iloc[0].get("Chromosome", ""))
                    for _, hit_row in grp.iterrows():
                        subj = str(hit_row.get("sseqid", "") or hit_row.get("subject", ""))
                        if clinvar_chrom and clinvar_chrom in subj:
                            nuclear_at_locus.add(aid)
                            break

    # Classification
    def assign_class(row):
        aid = row.get("allele_id_int")
        status = str(row.get("extraction_status", ""))

        # Déjà annoté comme référence mito dans le nom HGVS
        if status == "DIRECT_MITO_REF":
            return "DIRECT_MITO_REF"

        if pd.isna(aid):
            return "UNKNOWN"

        aid = int(aid)

        if aid not in mito_ids:
            return "NO_MITO_HIT"

        # A un hit mito
        if df_genome is None:
            return "MITO_HIT_ONLY"  # genome non testé

        if aid not in nuclear_ids:
            return "NOVEL_NUMT_CANDIDATE"  # ⭐ PAS dans le génome de référence

        if aid in nuclear_at_locus:
            return "CONFIRMED_NUMT"  # hit au locus ClinVar rapporté

        return "KNOWN_REFERENCE_NUMT"  # hit ailleurs dans hg38

    df["numt_class"] = df.apply(assign_class, axis=1)

    # Ajout des colonnes de meilleur hit mito
    for col in ["mito_pident", "mito_evalue", "mito_hit_tier", "mito_region",
                "mito_region_type", "mito_combined_score", "mito_n_hsps"]:
        df[col] = df["allele_id_int"].apply(
            lambda x: mito_best.get(int(x), {}).get(col) if not pd.isna(x) else None
        )

    # Estimation de l'âge d'insertion à partir du pident BLAST vs rCRS
    # Smart 2019 / HazkaniCovo 2003 : pident ≈ 99% → événement récent/de novo,
    # pident ≈ 80–85% → insertion ancienne.
    def _age_from_pident(pident):
        if pd.isna(pident):
            return None
        p = float(pident)
        if p >= 97:
            return "RECENT"       # de novo / très récent
        elif p >= 90:
            return "INTERMEDIATE"
        else:
            return "ANCIENT"      # insertion ancestrale

    df["insertion_age_estimate"] = df["mito_pident"].apply(_age_from_pident)

    # Normalisation de l'origine ClinVar → origin_group (germline / somatic / unknown)
    # OriginSimple est la colonne ClinVar pré-normalisée (valeurs: germline, somatic,
    # unknown, not applicable). On la préfère à Origin (qui contient de novo, maternal…).
    def _normalize_origin(row):
        val = str(row.get("OriginSimple", "") or "").strip().lower()
        if not val:
            val = str(row.get("Origin", "") or "").strip().lower()
        if val in ("germline", "de novo", "inherited", "maternal", "paternal",
                   "biparental", "uniparental"):
            return "germline"
        if val in ("somatic",):
            return "somatic"
        return "unknown"

    df["origin_group"] = df.apply(_normalize_origin, axis=1)

    # Marquage gènes de maladie
    df["is_disease_gene"] = df.get("GeneSymbol", pd.Series(dtype=str)).isin(DISEASE_GENES)

    df = df.drop(columns=["allele_id_int"], errors="ignore")
    return df


# ─── ÉTAPE 7 : analyse du mécanisme d'insertion ──────────────────────────────

def _load_mechanism_module():
    """Charge dynamiquement 06_mechanism.py sans importer son main()."""
    mod_path = SCRIPT_DIR / "06_mechanism.py"
    if not mod_path.exists():
        log(f"WARNING — 06_mechanism.py introuvable : {mod_path}")
        return None
    spec = importlib.util.spec_from_file_location("mechanism06", mod_path)
    mod  = importlib.util.module_from_spec(spec)
    try:
        spec.loader.exec_module(mod)
    except Exception as exc:
        log(f"WARNING — Impossible de charger 06_mechanism.py : {exc}")
        return None
    return mod


def run_mechanism_step(
    df_classified: pd.DataFrame,
    df_mito: pd.DataFrame,
    email: str | None,
) -> pd.DataFrame:
    """
    Lance l'analyse mécanistique sur les candidats NUMT (étape 7).

    Candidats analysés : NOVEL_NUMT_CANDIDATE, CONFIRMED_NUMT, DIRECT_MITO_REF.

    Pour chaque candidat :
      1. Construit un dict «doc» compatible avec 06_mechanism.analyse_candidate()
      2. Enrichit avec les coordonnées mito (sstart/send) pour la détection d'orientation
      3. Appelle analyse_candidate() → récupère microhomologie, répétitions, TSD, L1, mécanisme

    Retourne un DataFrame plat pour la feuille Excel «7_Mechanism».
    """
    mod = _load_mechanism_module()
    if mod is None:
        log("Étape 7 ignorée — module 06_mechanism.py non disponible.")
        return pd.DataFrame()

    analyse_candidate  = mod.analyse_candidate
    save_json_report   = mod.save_json_report
    save_tsv_summary   = mod.save_tsv_summary
    print_summary_table = mod.print_summary_table

    # Filtrer les candidats à analyser
    target_classes = {"NOVEL_NUMT_CANDIDATE", "CONFIRMED_NUMT", "DIRECT_MITO_REF"}
    df_candidates  = df_classified[df_classified["numt_class"].isin(target_classes)].copy()

    if df_candidates.empty:
        log("Étape 7 — Aucun candidat à analyser.")
        return pd.DataFrame()

    log(f"Étape 7 — Analyse mécanistique de {len(df_candidates):,} candidats ...")

    # Préparer un index AlleleID → meilleur hit mito (sstart/send pour orientation)
    mito_best_hit: dict[int, dict] = {}
    if not df_mito.empty and "allele_id" in df_mito.columns:
        for aid_raw, grp in df_mito.groupby("allele_id"):
            try:
                aid = int(float(aid_raw))
            except (ValueError, TypeError):
                continue
            # Meilleur HSP par combined_score
            if "combined_score" in grp.columns:
                best = grp.sort_values("combined_score", ascending=False).iloc[0]
            else:
                best = grp.iloc[0]
            mito_best_hit[aid] = {
                "sstart": best.get("sstart"),
                "send":   best.get("send"),
            }

    results: list[dict] = []
    total = len(df_candidates)

    for i, (_, row) in enumerate(df_candidates.iterrows(), 1):
        allele_id_raw = row.get("AlleleID", "")
        try:
            allele_id = int(float(allele_id_raw))
        except (ValueError, TypeError):
            allele_id = allele_id_raw

        log(f"  [{i}/{total}] AlleleID {allele_id} ({row.get('GeneSymbol', '?')}) ...")

        # Construction du dict doc pour analyse_candidate()
        doc = {
            "AlleleID":            allele_id,
            "GeneSymbol":          row.get("GeneSymbol", ""),
            "Chromosome":          str(row.get("Chromosome", "") or ""),
            "Start":               row.get("Start", ""),
            "Stop":                row.get("Stop", ""),
            "extracted_sequence":  row.get("extracted_sequence", ""),
        }

        # Enrichissement avec les coordonnées BLAST mito (pour orientation)
        if allele_id in mito_best_hit:
            doc["mito_blast"] = {
                "best_hit": mito_best_hit[allele_id]
            }

        result = analyse_candidate(
            doc,
            flank_size=500,
            min_microhomology=4,
            email=email or "numt_pipeline@example.com",
        )

        if result is not None:
            result["numt_class"] = row.get("numt_class", "")
            results.append(result)

    if not results:
        log("Étape 7 — Aucun résultat de mécanisme produit.")
        return pd.DataFrame()

    log(f"Étape 7 — {len(results):,} analyses terminées.")
    print_summary_table(results)

    # Sauvegarde JSON et TSV
    save_json_report(results, MECHANISM_JSON)
    save_tsv_summary(results, MECHANISM_TSV)

    # Aplatir les résultats en DataFrame pour Excel
    rows_flat = []
    for r in results:
        at  = r.get("at_content",   {})
        rep = r.get("repeat_context", {})
        rec = r.get("recombination", {})
        rows_flat.append({
            "AlleleID":                  r["allele_id"],
            "Gene":                      r["gene"],
            "numt_class":                r.get("numt_class", ""),
            "predicted_mechanism":       r["predicted_mechanism"],
            "mechanism_confidence":      r["mechanism_confidence"],
            "insertion_orientation":     r["insertion_orientation"],
            "microhomology_found":       r["microhomology"]["found"],
            "microhomology_length":      r["microhomology"]["length"],
            "microhomology_seq":         r["microhomology"]["sequence"],
            "tsd_found":                 r["tsd"]["found"],
            "tsd_sequence":              r["tsd"]["sequence"],
            "l1_site_found":             r["l1_site"]["found"],
            "l1_site_motif":             r["l1_site"].get("motif", ""),
            "direct_repeats_found":      r["direct_repeats"]["found"],
            "direct_repeats_seq":        r["direct_repeats"]["sequence"],
            # ── New columns (v2) ─────────────────────────────────────────
            "at_content_combined":       at.get("at_combined", ""),
            "dna_curvature_mean":        r.get("dna_curvature_mean", ""),
            "nearest_repeat_class":      rep.get("nearest_repeat_class", "N/A"),
            "nearest_repeat_dist_bp":    rep.get("nearest_repeat_dist_bp", -1),
            "l1_within_5kb":             rep.get("l1_within_5kb", False),
            "recombination_rate_cM_Mb":  rec.get("rate_cM_Mb",
                                             rec.get("recombination_rate_cM_Mb", "")),
            "in_recombination_hotspot":  rec.get("in_hotspot",
                                             rec.get("in_recombination_hotspot", False)),
            "interpretation":            r.get("interpretation", ""),
        })

    return pd.DataFrame(rows_flat)


# ─── Écriture Excel ───────────────────────────────────────────────────────────

# Couleurs par numt_class
NUMT_CLASS_COLORS = {
    "NOVEL_NUMT_CANDIDATE":  "90EE90",   # vert clair   ← priorité absolue
    "CONFIRMED_NUMT":        "FFFF99",   # jaune
    "KNOWN_REFERENCE_NUMT":  "FFD580",   # orange clair
    "DIRECT_MITO_REF":       "ADD8E6",   # bleu clair
    "MITO_HIT_ONLY":         "E0E0E0",   # gris clair
    "NO_MITO_HIT":           "FFFFFF",   # blanc
}

HIT_TIER_COLORS = {
    "STRONG_HIT":   "90EE90",
    "MODERATE_HIT": "FFFF99",
    "WEAK_HIT":     "FFD580",
}

MECHANISM_COLORS = {
    "TPRT":     "C8E6C9",   # vert pâle
    "NAHR":     "FFF9C4",   # jaune pâle
    "NHEJ":     "FFE0B2",   # orange pâle
    "MMEJ/MMBIR": "FFCCBC",   # saumon pâle
    "UNKNOWN":  "ECEFF1",   # gris très clair
}


def _autofit_columns(ws, max_width: int = 60) -> None:
    """Ajuste la largeur des colonnes au contenu."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _style_header_row(ws) -> None:
    """Met en gras et en fond gris la première ligne."""
    header_fill = PatternFill("solid", fgColor="BFBFBF")
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _colorize_column(ws, col_name: str, col_idx: int, color_map: dict) -> None:
    """Colorie les cellules d'une colonne selon une map {valeur: couleur hex}."""
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for cell in row:
            color = color_map.get(str(cell.value), None)
            if color:
                cell.fill = PatternFill("solid", fgColor=color)


def _write_statistics_sheet(
    ws,
    df_mechanism: pd.DataFrame,
    df_novel: pd.DataFrame,
    df_confirmed: pd.DataFrame | None = None,
) -> None:
    """
    Write the 8_Statistics sheet directly to an openpyxl worksheet.

    Produces nine tables with interpretation text, separated by blank rows:
      Table 1 — Mechanism distribution
      Table 2 — Clinical significance × mechanism
      Table 3 — mtDNA region × mechanism
      Table 4 — AT% by mechanism (mean ± SD)
      Table 5 — Mobile element proximity (UCSC RepeatMasker)
      Table 6 — Recombination hotspot overlap
      Table 7 — Insertion age estimate (pident vs rCRS)
      Table 8 — mtDNA source region × numt_class (Fisher's exact test)
      Table 9 — ClinVar origin (germline / somatic / unknown)

    Also appends a LIMITATIONS section citing key references.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt
    import numpy as np

    # ── Style helpers ─────────────────────────────────────────────────────────
    GRAY_FILL   = PatternFill("solid", fgColor="BFBFBF")
    YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")
    WHITE       = PatternFill("solid", fgColor="FFFFFF")

    # Try to import mechanism interpretations from 06_mechanism module
    _mod = _load_mechanism_module()
    MECH_INTERP = getattr(_mod, "MECHANISM_INTERPRETATIONS", {}) if _mod else {}

    row = [1]  # mutable counter

    def _cell(r, c, value="", bold=False, italic=False, fill=None, wrap=False,
              size=11):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = Font(bold=bold, italic=italic, size=size)
        if fill:
            cell.fill = fill
        if wrap:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        return cell

    def write_section_header(title):
        _cell(row[0], 1, title, bold=True, size=12)
        ws.row_dimensions[row[0]].height = 20
        row[0] += 1

    def write_col_headers(*cols):
        for c, col in enumerate(cols, 1):
            _cell(row[0], c, col, bold=True, fill=GRAY_FILL)
        row[0] += 1

    def write_data_row(*vals, highlight=False):
        fill = YELLOW_FILL if highlight else None
        for c, v in enumerate(vals, 1):
            _cell(row[0], c, v, fill=fill)
        row[0] += 1

    def write_note(text):
        _cell(row[0], 1, text, italic=True, wrap=True)
        ws.row_dimensions[row[0]].height = 72
        row[0] += 1

    def blank(n=1):
        row[0] += n

    # Wide columns for text
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 80

    # ── Title ─────────────────────────────────────────────────────────────────
    _cell(1, 1,
          f"MECHANISM ANALYSIS STATISTICS — NOVEL NUMT CANDIDATES",
          bold=True, size=14)
    row[0] = 2
    _cell(row[0], 1,
          f"Generated {_dt.now().strftime('%Y-%m-%d')}  |  "
          "Analysis uses hg38 reference flanking sequences (500 bp)  |  "
          "n = 21 candidates total (20 NOVEL_NUMT_CANDIDATE + 1 CONFIRMED_NUMT)",
          italic=True)
    row[0] += 1
    blank()

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 1: Mechanism distribution
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 1 — Mechanism Distribution")
    write_col_headers("Mechanism", "n", "%", "MEDIUM+ confidence", "Short definition")

    total = len(df_mechanism)
    mech_counts = df_mechanism["predicted_mechanism"].value_counts()
    if "mechanism_confidence" in df_mechanism.columns:
        med_plus = (
            df_mechanism[df_mechanism["mechanism_confidence"].isin(["MEDIUM", "HIGH"])]
            ["predicted_mechanism"].value_counts()
        )
    else:
        med_plus = {}

    for mech in ["TPRT", "NAHR", "MMEJ/MMBIR", "NHEJ", "UNKNOWN"]:
        n = int(mech_counts.get(mech, 0))
        if n == 0:
            continue
        pct = f"{100*n/total:.1f}%"
        mp  = int(med_plus.get(mech, 0))
        defn = MECH_INTERP.get(mech, "")
        # Truncate for cell display
        defn_short = defn[:120] + "…" if len(defn) > 120 else defn
        write_data_row(mech, n, pct, mp, defn_short)
    write_data_row("Total", total, "100%", "", "")

    blank()
    full_interp = "\n\n".join(
        f"{m}: {MECH_INTERP[m]}"
        for m in mech_counts.index
        if m in MECH_INTERP
    )
    write_note(
        "Interpretation: " + full_interp +
        "\n\nReferences: Tsuji et al. (2012) Nucleic Acids Res 40:9272–9282; "
        "Xue et al. (2023) Biomolecules 13:1763; "
        "Hazkani-Covo et al. (2003) J Mol Evol 57:343–357; "
        "Bíró et al. (2024)."
    )
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 2: Clinical significance × mechanism
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 2 — Clinical Significance × Mechanism")

    clinsig_col = None
    for col in ("ClinicalSignificance", "clinical_significance", "ClinSig"):
        if col in df_mechanism.columns:
            clinsig_col = col
            break
    # Fall back to df_novel if not in df_mechanism
    if clinsig_col is None and not df_novel.empty:
        for col in ("ClinicalSignificance", "ClinSig"):
            if col in df_novel.columns:
                # Merge ClinSig into mechanism df
                df_cs = df_novel[["AlleleID", col]].copy()
                df_cs["AlleleID"] = df_cs["AlleleID"].astype(str)
                df_m2 = df_mechanism.copy()
                df_m2["AlleleID"] = df_m2["AlleleID"].astype(str)
                df_m2 = df_m2.merge(df_cs, on="AlleleID", how="left")
                clinsig_col = col
                df_mechanism = df_m2
                break

    if clinsig_col and clinsig_col in df_mechanism.columns:
        mechs_present = [m for m in ["TPRT", "NAHR", "MMEJ/MMBIR", "NHEJ", "UNKNOWN"]
                         if m in mech_counts.index]
        write_col_headers("ClinicalSignificance", *mechs_present, "Total")
        ct = pd.crosstab(df_mechanism[clinsig_col], df_mechanism["predicted_mechanism"])
        for sig, row_data in ct.iterrows():
            vals = [sig] + [int(row_data.get(m, 0)) for m in mechs_present]
            vals.append(int(row_data.sum()))
            write_data_row(*vals)
        # Total row
        totals = ["Total"] + [int(ct[m].sum()) if m in ct.columns else 0
                               for m in mechs_present]
        totals.append(int(ct.values.sum()))
        write_data_row(*totals)
    else:
        _cell(row[0], 1, "ClinicalSignificance column not available in this run.", italic=True)
        row[0] += 1
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 3: mtDNA region × mechanism
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 3 — mtDNA Region × Mechanism")

    mito_col = None
    for col in ("mito_region", "mito_hit_region"):
        if col in df_mechanism.columns:
            mito_col = col
            break
    if mito_col is None and not df_novel.empty:
        for col in ("mito_region",):
            if col in df_novel.columns:
                df_mr = df_novel[["AlleleID", col]].copy()
                df_mr["AlleleID"] = df_mr["AlleleID"].astype(str)
                df_m3 = df_mechanism.copy()
                df_m3["AlleleID"] = df_m3["AlleleID"].astype(str)
                df_m3 = df_m3.merge(df_mr, on="AlleleID", how="left")
                mito_col = col
                df_mechanism = df_m3
                break

    if mito_col and mito_col in df_mechanism.columns:
        mechs_present = [m for m in ["TPRT", "NAHR", "MMEJ/MMBIR", "NHEJ", "UNKNOWN"]
                         if m in mech_counts.index]
        write_col_headers("mtDNA Region", *mechs_present, "Total")
        ct3 = pd.crosstab(df_mechanism[mito_col], df_mechanism["predicted_mechanism"])
        for region, row_data in ct3.iterrows():
            vals = [str(region)] + [int(row_data.get(m, 0)) for m in mechs_present]
            vals.append(int(row_data.sum()))
            write_data_row(*vals)
        totals3 = ["Total"] + [int(ct3[m].sum()) if m in ct3.columns else 0
                                for m in mechs_present]
        totals3.append(int(ct3.values.sum()))
        write_data_row(*totals3)
    else:
        _cell(row[0], 1, "mito_region column not available — merge with NOVEL_CANDIDATES sheet.",
              italic=True)
        row[0] += 1
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 4: AT% by mechanism
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 4 — AT% by Mechanism (mean ± SD)")

    if "at_content_combined" in df_mechanism.columns:
        df_at = df_mechanism[["predicted_mechanism", "at_content_combined"]].copy()
        df_at["at_content_combined"] = pd.to_numeric(
            df_at["at_content_combined"], errors="coerce"
        )
        write_col_headers("Mechanism", "n", "Mean AT%", "SD AT%", "Min AT%", "Max AT%")
        for mech in mech_counts.index:
            grp = df_at[df_at["predicted_mechanism"] == mech]["at_content_combined"].dropna()
            if len(grp) == 0:
                continue
            write_data_row(
                mech, len(grp),
                f"{grp.mean():.1f}%",
                f"{grp.std():.1f}%",
                f"{grp.min():.1f}%",
                f"{grp.max():.1f}%",
            )
        write_note(
            "High AT% (>60%) flanking sequences are a hallmark of NUMT insertion sites "
            "(Tsuji et al. 2012). AT-rich regions are associated with open chromatin and "
            "higher rates of double-strand breaks."
        )
    else:
        _cell(row[0], 1,
              "at_content_combined column not present — re-run pipeline with updated "
              "06_mechanism.py to populate this column.", italic=True)
        row[0] += 1
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 5: Mobile element proximity
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 5 — Mobile Element Proximity (UCSC RepeatMasker ±5 kb)")

    if "nearest_repeat_class" in df_mechanism.columns:
        df_rep = df_mechanism[["predicted_mechanism", "nearest_repeat_class",
                                "nearest_repeat_dist_bp", "l1_within_5kb"]].copy()
        # Distribution of repeat classes
        rep_counts = df_rep["nearest_repeat_class"].value_counts()
        write_col_headers("Repeat Class", "n", "% of all candidates",
                          "Mean distance (bp)", "Notes")
        for rep_class, cnt in rep_counts.items():
            sub = df_rep[df_rep["nearest_repeat_class"] == rep_class]
            dist_col = pd.to_numeric(sub["nearest_repeat_dist_bp"], errors="coerce")
            mean_dist = f"{dist_col[dist_col >= 0].mean():.0f}" if (dist_col >= 0).any() else "N/A"
            write_data_row(
                rep_class, int(cnt),
                f"{100*cnt/total:.1f}%",
                mean_dist, "",
            )
        blank()
        # L1 summary
        if "l1_within_5kb" in df_mechanism.columns:
            l1_n = int(df_mechanism["l1_within_5kb"].sum())
            write_col_headers("L1 within 5 kb", "n", "% of candidates")
            write_data_row("Yes", l1_n, f"{100*l1_n/total:.1f}%")
            write_data_row("No",  total - l1_n, f"{100*(total-l1_n)/total:.1f}%")
        write_note(
            "Mobile element proximity: UCSC RepeatMasker (rmsk) track queried for each "
            "insertion site (window = ±5 kb). LINE-1 (L1) elements within 5 kb are "
            "relevant to TPRT mechanism (Xue et al. 2023). Repeat-rich flanking regions "
            "are also associated with NAHR (Bíró et al. 2024)."
        )
    else:
        _cell(row[0], 1,
              "nearest_repeat_class column not present — requires updated 06_mechanism.py "
              "with UCSC API queries.", italic=True)
        row[0] += 1
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 6: Recombination hotspot overlap
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 6 — Recombination Rate / Hotspot Overlap")

    if "recombination_rate_cM_Mb" in df_mechanism.columns:
        df_rec = df_mechanism[["predicted_mechanism", "recombination_rate_cM_Mb",
                                "in_recombination_hotspot"]].copy()
        df_rec["recombination_rate_cM_Mb"] = pd.to_numeric(
            df_rec["recombination_rate_cM_Mb"], errors="coerce"
        )
        write_col_headers("Metric", "Value", "Notes")
        rate_col = df_rec["recombination_rate_cM_Mb"].dropna()
        write_data_row("Mean recombination rate (all)", f"{rate_col.mean():.2f} cM/Mb", "")
        write_data_row("Median recombination rate", f"{rate_col.median():.2f} cM/Mb", "")
        write_data_row("Max recombination rate", f"{rate_col.max():.2f} cM/Mb", "")

        if "in_recombination_hotspot" in df_mechanism.columns:
            hotspot_n = int(df_mechanism["in_recombination_hotspot"].sum())
            write_data_row("Candidates in hotspot (>10 cM/Mb)",
                           f"{hotspot_n}/{total}  ({100*hotspot_n/total:.1f}%)", "")
        blank()
        # By mechanism
        write_col_headers("Mechanism", "Mean rate (cM/Mb)", "In hotspot (n)")
        for mech in mech_counts.index:
            grp = df_rec[df_rec["predicted_mechanism"] == mech]
            r_vals = pd.to_numeric(grp["recombination_rate_cM_Mb"], errors="coerce").dropna()
            hs_n   = int(grp["in_recombination_hotspot"].sum()) if "in_recombination_hotspot" \
                        in grp.columns else 0
            mean_r = f"{r_vals.mean():.2f}" if len(r_vals) > 0 else "N/A"
            write_data_row(mech, mean_r, hs_n)
        write_note(
            "Recombination rate from UCSC recombRate track (deCODE/hg38). "
            "PRDM9 binding sites are not directly available as a UCSC track; "
            "recombination rate >10 cM/Mb is used as a proxy for meiotic hotspot activity. "
            "True PRDM9 analysis would require ChIP-seq data (not available here)."
        )
    else:
        _cell(row[0], 1,
              "recombination_rate_cM_Mb column not present — requires updated "
              "06_mechanism.py with UCSC API queries.", italic=True)
        row[0] += 1
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 7: Insertion age estimate (pident vs rCRS)
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 7 — Insertion Age Estimate (pident vs rCRS)")

    age_col = "insertion_age_estimate"
    pident_col = "mito_pident"
    df_age_src = df_novel if not df_novel.empty else df_mechanism
    if age_col in df_age_src.columns:
        age_counts = df_age_src[age_col].value_counts()
        n_age = len(df_age_src)
        write_col_headers("Age estimate", "pident range", "n", "%", "Interpretation")
        for label, prange, interp in [
            ("RECENT",       "≥97%",     "de novo / very recent insertion"),
            ("INTERMEDIATE", "90–97%",   "post-divergence, not ancient"),
            ("ANCIENT",      "<90%",     "ancestral NUMT"),
        ]:
            n = int(age_counts.get(label, 0))
            write_data_row(label, prange, n,
                           f"{100*n/n_age:.1f}%" if n_age else "—", interp)
        blank()
        if pident_col in df_age_src.columns:
            pident_vals = pd.to_numeric(df_age_src[pident_col], errors="coerce").dropna()
            if len(pident_vals) > 0:
                write_col_headers("pident stats", "Value")
                write_data_row("Mean pident", f"{pident_vals.mean():.1f}%")
                write_data_row("Min pident",  f"{pident_vals.min():.1f}%")
                write_data_row("Max pident",  f"{pident_vals.max():.1f}%")
        write_note(
            "Insertion age estimated from BLAST pident vs rCRS (NC_012920.1). "
            "RECENT (pident ≥97%) candidates are prioritised as likely de novo events. "
            "Rationale: sequence divergence from rCRS correlates with insertion age "
            "(Smart et al. 2019; Hazkani-Covo et al. 2003). Haplogroup bias possible: "
            "rCRS represents haplogroup H; non-H insertions may appear falsely ancient."
        )
    else:
        _cell(row[0], 1,
              "insertion_age_estimate column not present — re-run classification step.",
              italic=True)
        row[0] += 1
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 8: mtDNA source region × numt_class (Fisher's exact test)
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 8 — mtDNA Source Region × numt_class (D-loop enrichment test)")

    _region_col = "mito_region"
    _class_col  = "numt_class"
    _df_conf    = df_confirmed if isinstance(df_confirmed, pd.DataFrame) and not df_confirmed.empty \
                  else pd.DataFrame()

    _novel_has_region    = not df_novel.empty and _region_col in df_novel.columns
    _conf_has_region     = not _df_conf.empty  and _region_col in _df_conf.columns and \
                            _class_col in _df_conf.columns

    if _novel_has_region:
        # Build combined DataFrame with numt_class labels
        _parts = []
        if _novel_has_region:
            _n = df_novel[[_region_col]].copy()
            _n[_class_col] = "NOVEL_NUMT_CANDIDATE"
            _parts.append(_n)
        if _conf_has_region:
            _c = _df_conf[[_region_col, _class_col]].copy()
            _parts.append(_c)

        _df_all = pd.concat(_parts, ignore_index=True)

        # All region × class crosstab
        _classes_present = [c for c in
                            ["NOVEL_NUMT_CANDIDATE", "CONFIRMED_NUMT", "KNOWN_REFERENCE_NUMT"]
                            if c in _df_all[_class_col].values]
        write_col_headers("mtDNA region", *_classes_present, "Total")
        _ct8 = pd.crosstab(_df_all[_region_col], _df_all[_class_col])
        for region, rdata in _ct8.iterrows():
            vals = [str(region)] + [int(rdata.get(c, 0)) for c in _classes_present]
            vals.append(int(rdata.sum()))
            write_data_row(*vals)
        _tot8 = ["Total"] + [int(_ct8[c].sum()) if c in _ct8.columns else 0
                              for c in _classes_present]
        _tot8.append(int(_ct8.values.sum()))
        write_data_row(*_tot8)
        blank()

        # Fisher's exact test: D-loop vs non-D-loop × NOVEL vs CONFIRMED+KNOWN_REF
        _novel_regions  = df_novel[_region_col].fillna("intergenic")
        _n_dloop_novel  = int((_novel_regions == "D-loop").sum())
        _n_other_novel  = int((_novel_regions != "D-loop").sum())

        _conf_regions   = _df_conf[_region_col].fillna("intergenic") if _conf_has_region \
                          else pd.Series([], dtype=str)
        _n_dloop_conf   = int((_conf_regions == "D-loop").sum())
        _n_other_conf   = int((_conf_regions != "D-loop").sum())

        _contingency = [[_n_dloop_novel, _n_other_novel],
                        [_n_dloop_conf,  _n_other_conf]]

        write_col_headers("Fisher 2×2", "D-loop", "Non-D-loop", "Total")
        write_data_row("NOVEL_NUMT_CANDIDATE",
                       _n_dloop_novel, _n_other_novel, _n_dloop_novel + _n_other_novel)
        write_data_row("CONFIRMED+KNOWN_REF",
                       _n_dloop_conf, _n_other_conf, _n_dloop_conf + _n_other_conf)
        blank()

        # Try Fisher's exact via scipy
        _fisher_text = "Fisher's exact test: scipy not available — install scipy for p-value."
        try:
            from scipy.stats import fisher_exact
            _oddsratio, _pvalue = fisher_exact(_contingency, alternative="greater")
            if _n_dloop_conf + _n_other_conf < 2:
                _fisher_text = (
                    f"Fisher's exact test (D-loop enrichment in NOVEL): "
                    f"CONFIRMED group too small (n={_n_dloop_conf + _n_other_conf}) for "
                    f"meaningful comparison — OR={_oddsratio:.2f}, p={_pvalue:.4f} "
                    f"(interpret with caution)."
                )
            else:
                _fisher_text = (
                    f"Fisher's exact test (one-sided, D-loop enrichment in NOVEL vs CONFIRMED): "
                    f"OR={_oddsratio:.2f}, p={_pvalue:.4f}. "
                    + ("Significant at α=0.05 — D-loop is enriched among NOVEL candidates."
                       if _pvalue < 0.05 else
                       "Not significant at α=0.05 — D-loop enrichment hypothesis not supported.")
                )
        except ImportError:
            pass
        except Exception as _exc:
            _fisher_text = f"Fisher's exact test error: {_exc}"

        write_note(
            _fisher_text +
            "\n\nInterpretation: D-loop (control region) variants may be overrepresented "
            "among NOVEL_NUMT_CANDIDATE because the control region lacks selective pressure, "
            "making de novo insertions from this region more likely to reach fixation or be "
            "transmitted (Hazkani-Covo et al. 2003; Mourier et al. 2008). "
            "If p>0.05, the D-loop hypothesis is not supported by ClinVar data alone, "
            "possibly due to ascertainment bias (ClinVar captures protein-coding gene variants)."
        )
    else:
        _cell(row[0], 1,
              "mito_region column not available in NOVEL_CANDIDATES — "
              "re-run classify_variants() step.", italic=True)
        row[0] += 1
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 9: ClinVar origin — germline vs somatic
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("TABLE 9 — ClinVar Origin (germline / somatic / unknown)")

    _og_col = "origin_group"
    # Combine all mito-hit candidates: novel + confirmed (both carry origin_group)
    _parts9 = []
    if not df_novel.empty and _og_col in df_novel.columns:
        _n9 = df_novel[[_og_col, _class_col]].copy()
        _parts9.append(_n9)
    if isinstance(df_confirmed, pd.DataFrame) and not df_confirmed.empty \
            and _og_col in df_confirmed.columns and _class_col in df_confirmed.columns:
        _parts9.append(df_confirmed[[_og_col, _class_col]].copy())

    if _parts9:
        _df9 = pd.concat(_parts9, ignore_index=True)
        _classes9 = [c for c in
                     ["NOVEL_NUMT_CANDIDATE", "CONFIRMED_NUMT", "KNOWN_REFERENCE_NUMT"]
                     if c in _df9[_class_col].values]

        # Full origin × class crosstab
        write_col_headers("Origin", *_classes9, "Total")
        _ct9 = pd.crosstab(_df9[_og_col], _df9[_class_col])
        for origin, rdata in _ct9.iterrows():
            vals9 = [str(origin)] + [int(rdata.get(c, 0)) for c in _classes9]
            vals9.append(int(rdata.sum()))
            # Highlight somatic rows
            write_data_row(*vals9, highlight=(str(origin) == "somatic"))
        _tot9 = ["Total"] + [int(_ct9[c].sum()) if c in _ct9.columns else 0
                              for c in _classes9]
        _tot9.append(int(_ct9.values.sum()))
        write_data_row(*_tot9)
        blank()

        # Somatic-specific annotation
        _n_somatic = int((_df9[_og_col] == "somatic").sum())
        if _n_somatic > 0:
            write_note(
                f"WARNING — {_n_somatic} somatic candidate(s) identified. "
                "Somatic NUMTs are distinct from germline events: "
                "(1) no inherited recurrence risk; "
                "(2) may represent numtogenesis — active mtDNA integration during tumour "
                "evolution (Singh et al. 2017 Mitochondrion); "
                "(3) verify in COSMIC / TCGA before clinical reporting. "
                "These should NOT be classified using germline ACMG criteria."
            )
        else:
            write_note(
                "All mito-hit candidates are of germline origin (or unknown). "
                "No somatic numtogenesis event identified in this ClinVar dataset. "
                "Note: ClinVar is predominantly germline — somatic variants may be "
                "under-represented relative to databases such as COSMIC or cBioPortal."
            )
    else:
        _cell(row[0], 1,
              "origin_group column not available — re-run classify_variants() step.",
              italic=True)
        row[0] += 1
    blank(2)

    # ─────────────────────────────────────────────────────────────────────────
    # LIMITATIONS
    # ─────────────────────────────────────────────────────────────────────────
    write_section_header("LIMITATIONS")
    limitations = [
        "1. Reference genome only — flanking sequences are from hg38, not patient-specific "
        "alleles. Patient DNA may differ at the insertion site.",
        "2. ClinVar breakpoints approximate — HGVS positions may be ±1 bp off the true "
        "junction, making TSD and microhomology detection approximate.",
        "3. Short direct repeats (10–15 bp) are common by chance — NAHR calls based on "
        "short repeats remain LOW/MEDIUM confidence and require experimental validation.",
        "4. PRDM9/hotspot proxy — true PRDM9 ChIP-seq data not used; recombination rate "
        "from deCODE is a population-level proxy only.",
        "5. No poly-A detection in any candidate — TPRT cannot be excluded without patient "
        "RNA or long-read sequencing of the junction.",
        "6. UCSC API — queries may fail if network is unavailable; missing values are "
        "reported as N/A or 0.0.",
        "7. MMEJ and MMBIR are mechanistically distinct but leave identical junction "
        "microhomology signatures detectable by sequence analysis. Their distinction "
        "requires patient-derived long-read sequencing to resolve templated insertions "
        "or complex junction architecture (Hastings et al. 2009; Bíró et al. 2024).",
        "8. Insertion age estimate (pident vs rCRS) assumes haplogroup H: rCRS represents "
        "haplogroup H. Insertions from patients with distant haplogroups (L, M, N…) may "
        "show artificially lower pident, leading to overestimation of insertion age. "
        "A pan-mitogenome BLAST database would improve sensitivity.",
    ]
    for lim in limitations:
        write_note(lim)

    # Final autofit
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        if col_letter in ("A", "E"):
            continue  # already set wide
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)


def write_excel(
    sheets: dict[str, pd.DataFrame],
    output_path: Path,
) -> None:
    """
    Écrit le classeur Excel avec une feuille par étape.

    sheets : dict { nom_feuille → DataFrame }
    """
    if not OPENPYXL_AVAILABLE:
        log("openpyxl non installé — sauvegarde en TSV de secours.")
        for name, df in sheets.items():
            safe_name = name.replace(" ", "_").replace("/", "_")
            tsv_path  = output_path.with_name(f"{safe_name}.tsv")
            df.to_csv(tsv_path, sep="\t", index=False)
            log(f"  Sauvegardé : {tsv_path}")
        return

    log(f"Écriture du fichier Excel : {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            # Special handling for the statistics sheet
            if sheet_name == "8_Statistics" and isinstance(df, tuple):
                if len(df) == 3:
                    df_mech, df_nov, df_conf = df
                else:
                    df_mech, df_nov = df
                    df_conf = pd.DataFrame()
                wb = writer.book
                ws = wb.create_sheet(sheet_name[:31])
                try:
                    _write_statistics_sheet(ws, df_mech, df_nov, df_conf)
                except Exception as exc:
                    log(f"WARNING — statistics sheet error: {exc}")
                continue

            if not isinstance(df, pd.DataFrame) or df.empty:
                # Créer quand même la feuille avec un message
                df_empty = pd.DataFrame({"info": ["Aucune donnée pour cette étape."]})
                df_empty.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                continue

            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            ws = writer.sheets[sheet_name[:31]]
            _style_header_row(ws)
            _autofit_columns(ws)

            # Colorier la colonne numt_class si elle existe
            if "numt_class" in df.columns:
                col_idx = df.columns.get_loc("numt_class") + 1
                _colorize_column(ws, "numt_class", col_idx, NUMT_CLASS_COLORS)

            # Colorier la colonne hit_tier si elle existe
            if "hit_tier" in df.columns:
                col_idx = df.columns.get_loc("hit_tier") + 1
                _colorize_column(ws, "hit_tier", col_idx, HIT_TIER_COLORS)

            # Colorier la colonne predicted_mechanism si elle existe
            if "predicted_mechanism" in df.columns:
                col_idx = df.columns.get_loc("predicted_mechanism") + 1
                _colorize_column(ws, "predicted_mechanism", col_idx, MECHANISM_COLORS)

            # Freeze la première ligne (en-tête)
            ws.freeze_panes = "A2"

    log(f"Excel écrit : {output_path}")
    log(f"Feuilles : {list(sheets.keys())}")


# ─── Orchestrateur principal ──────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pipeline NUMT sans MongoDB — sortie Excel multi-feuilles.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--from-step", type=int, choices=[1, 2, 3, 4, 5, 6, 7], default=1, metavar="N",
        help="Reprendre depuis l'étape N (1=tout refaire, 3=TSV déjà produits, 7=mécanisme seul)",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Forcer le re-téléchargement ClinVar (passé à l'étape 1)",
    )
    parser.add_argument(
        "--email", type=str, default=None, metavar="EMAIL",
        help="Email NCBI Entrez (requis pour le BLAST remote et le téléchargement rCRS)",
    )
    parser.add_argument(
        "--skip-genome", action="store_true",
        help=(
            "Sauter le BLAST vs génome humain. "
            "Les variants avec hit mito seront classés MITO_HIT_ONLY."
        ),
    )
    parser.add_argument(
        "--genome-db", type=Path, default=None, metavar="PATH",
        help=(
            "Chemin vers la base BLAST hg38 locale (sans extension). "
            "Si absent et --skip-genome non fourni, utilise le mode remote NCBI."
        ),
    )
    parser.add_argument(
        "--validate-last", action="store_true",
        help=(
            "Lancer LAST (last-train + lastal | last-split) en parallèle de BLAST "
            "pour validation. Ajoute les feuilles 4b_LAST_Mito et LAST_vs_BLAST."
        ),
    )
    parser.add_argument(
        "--output", type=Path, default=OUTPUT_EXCEL,
        help=f"Chemin du fichier Excel de sortie (défaut: {OUTPUT_EXCEL.name})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    log("=" * 65)
    log("NUMT Pipeline Excel — sans MongoDB")
    log("=" * 65)
    log(f"Reprise depuis l'étape : {args.from_step}")

    sheets: dict[str, pd.DataFrame] = {}

    # ── ÉTAPE 1 : ClinVar fetch ──────────────────────────────────────────
    if args.from_step <= 1:
        log("")
        log("─── Étape 1 : ClinVar Download & Filter ───")
        extra = ["--force"] if args.force else []
        ok = run_step("01_clinvar_fetch.py", extra)
        if not ok:
            log("FATAL — Étape 1 a échoué. Vérifiez la connexion réseau.")
            sys.exit(1)

    if not CLINVAR_RAW_TSV.exists():
        log(f"FATAL — Fichier attendu introuvable : {CLINVAR_RAW_TSV}")
        log("Lancez depuis l'étape 1 : --from-step 1")
        sys.exit(1)

    df_raw = pd.read_csv(CLINVAR_RAW_TSV, sep="\t", dtype=str, low_memory=False)
    log(f"Étape 1 → {len(df_raw):,} variants chargés")
    sheets["1_ClinVar_Raw"] = df_raw

    # ── ÉTAPE 2 : extraction de séquences ───────────────────────────────
    if args.from_step <= 2:
        log("")
        log("─── Étape 2 : Sequence Extraction (HGVS) ───")
        ok = run_step("02_sequence_extract.py")
        if not ok:
            log("FATAL — Étape 2 a échoué.")
            sys.exit(1)

    if not SEQUENCES_TSV.exists():
        log(f"FATAL — Fichier attendu introuvable : {SEQUENCES_TSV}")
        sys.exit(1)

    df_seq = pd.read_csv(SEQUENCES_TSV, sep="\t", dtype=str, low_memory=False)
    log(f"Étape 2 → {len(df_seq):,} variants avec extraction_status")
    sheets["2_Sequences"] = df_seq

    # ── ÉTAPE 3 : dédoublonnage (remplace MongoDB) ───────────────────────
    log("")
    log("─── Étape 3 : Dédoublonnage AlleleID (GRCh38 préféré) ───")
    df_dedup = dedup_by_allele(df_seq)
    sheets["3_Deduplicated"] = df_dedup

    # Statistiques extraction_status
    status_counts = df_dedup["extraction_status"].value_counts()
    for status, count in status_counts.items():
        log(f"  {status:<30} {count:>6,}  ({count/len(df_dedup)*100:.1f}%)")

    # ── ÉTAPE 4 : BLAST vs mitochondrie ─────────────────────────────────
    if args.from_step <= 4:
        log("")
        log("─── Étape 4 : BLAST vs Mitochondrie ───")

        rcrs_seq = ensure_rcrs(email=args.email)
        db_path  = ensure_blast_db()

        if db_path is None:
            log("FATAL — Impossible de construire la base BLAST mito.")
            sys.exit(1)

        n_queries = build_query_fasta(df_dedup, rcrs_seq=rcrs_seq)
        if n_queries == 0:
            log("FATAL — Aucune séquence dans le FASTA de requête.")
            sys.exit(1)

        df_mito = run_blast_mito(db_path)
    else:
        # Charger résultats existants
        if BLAST_MITO_TSV.exists() and BLAST_MITO_TSV.stat().st_size > 0:
            df_mito = pd.read_csv(BLAST_MITO_TSV, sep="\t", header=None, names=BLAST_COLS, dtype=str)
            for col in ["pident", "length", "evalue", "bitscore", "qlen", "slen",
                        "qstart", "qend", "sstart", "send"]:
                df_mito[col] = pd.to_numeric(df_mito[col], errors="coerce")
            df_mito["allele_id"] = df_mito["qseqid"].apply(
                lambda x: int(x.split("|")[0]) if "|" in str(x) and x.split("|")[0].isdigit() else None
            )
            df_mito["gene_symbol"] = df_mito["qseqid"].apply(
                lambda x: x.split("|")[1] if "|" in str(x) and len(x.split("|")) > 1 else None
            )
            df_mito["hit_tier"] = df_mito.apply(classify_hit, axis=1)
            regions = df_mito.apply(
                lambda r: annotate_region(int(r["sstart"]), int(r["send"])), axis=1
            )
            df_mito["mtdna_region"]      = regions.apply(lambda x: x[0])
            df_mito["mtdna_region_type"] = regions.apply(lambda x: x[1])
            df_mito["query_coverage"]  = (df_mito["length"] / df_mito["qlen"]) * 100.0
            df_mito["combined_score"]  = df_mito["pident"] * (df_mito["length"] / df_mito["qlen"]) / 100.0
            log(f"Étape 4 — résultats existants chargés : {len(df_mito):,} HSPs")
        else:
            log("WARNING — Résultats BLAST mito introuvables.")
            df_mito = pd.DataFrame()

    sheets["4_BLAST_Mito"] = df_mito

    n_mito_variants = df_mito["allele_id"].nunique() if not df_mito.empty else 0
    log(f"Étape 4 → {n_mito_variants:,} variants avec hit mitochondrial")

    # ── ÉTAPE 4b (optionnelle) : LAST validation vs mitochondrie ────────
    if args.validate_last and QUERY_FASTA.exists():
        log("")
        log("─── Étape 4b : LAST validation vs Mitochondrie ───")

        from last_validation.last_mito import run_last_mito
        from last_validation.compare_blast_last import compare as compare_blast_last

        df_last = run_last_mito(
            query=QUERY_FASTA,
            db_fasta=MITO_FASTA,
            workdir=LAST_WORKDIR,
            output=LAST_MITO_TSV,
        )
        sheets["4b_LAST_Mito"] = df_last

        n_last_variants = df_last["allele_id"].nunique() if not df_last.empty else 0
        log(f"Étape 4b → {n_last_variants:,} variants avec hit LAST")

        # Comparison
        if not df_last.empty and BLAST_MITO_TSV.exists():
            df_compare = compare_blast_last(
                blast_tsv=BLAST_MITO_TSV,
                last_tsv=LAST_MITO_TSV,
                output_tsv=LAST_COMPARE_TSV,
                output_report=LAST_REPORT_TXT,
            )
            sheets["LAST_vs_BLAST"] = df_compare

    # ── ÉTAPE 5 : BLAST vs génome humain ────────────────────────────────
    log("")
    log("─── Étape 5 : BLAST vs Génome Humain ───")
    df_genome: pd.DataFrame | None = None

    if args.skip_genome:
        log("Étape 5 ignorée (--skip-genome).")
        log("Les variants avec hit mito seront classés MITO_HIT_ONLY.")
        sheets["5_BLAST_Genome"] = pd.DataFrame(
            {"info": ["Étape ignorée. Utilisez --genome-db ou relancez sans --skip-genome."]}
        )
    elif args.from_step <= 5:
        if df_mito.empty:
            log("Pas de hits mito → BLAST génome ignoré.")
            sheets["5_BLAST_Genome"] = pd.DataFrame(
                {"info": ["Aucun hit mitochondrial — BLAST génome non nécessaire."]}
            )
        else:
            genome_fasta = build_genome_fasta(df_mito, df_dedup)

            if args.genome_db is not None:
                df_genome = run_blast_genome_local(genome_fasta, args.genome_db)
            else:
                if args.email is None:
                    log("WARNING — --email requis pour le BLAST remote NCBI.")
                    log("  Ajoutez --email votre@email.com ou --skip-genome")
                    df_genome = None
                    sheets["5_BLAST_Genome"] = pd.DataFrame(
                        {"info": ["Email manquant pour BLAST remote. Fournissez --email."]}
                    )
                else:
                    log("Mode remote NCBI — peut être lent (10s/requête)")
                    df_genome = run_blast_genome_remote(genome_fasta, args.email)

            if df_genome is not None:
                sheets["5_BLAST_Genome"] = df_genome
    else:
        # Charger résultats existants
        if BLAST_GENOME_TSV.exists() and BLAST_GENOME_TSV.stat().st_size > 0:
            df_genome = pd.read_csv(BLAST_GENOME_TSV, sep="\t", dtype=str)
            log(f"Étape 5 — résultats existants chargés : {len(df_genome):,} HSPs")
            sheets["5_BLAST_Genome"] = df_genome
        else:
            log("WARNING — Résultats BLAST génome introuvables, classification partielle.")

    # ── ÉTAPE 6 : classification finale ─────────────────────────────────
    log("")
    log("─── Étape 6 : Classification NUMT ───")

    df_classified = classify_variants(df_dedup, df_mito, df_genome)
    sheets["6_Classification"] = df_classified

    class_counts = df_classified["numt_class"].value_counts()
    for cls, count in class_counts.items():
        marker = "  ⭐" if cls == "NOVEL_NUMT_CANDIDATE" else ""
        log(f"  {cls:<30} {count:>6,}{marker}")

    # ── Feuilles filtrées ────────────────────────────────────────────────
    _age_order = {"RECENT": 0, "INTERMEDIATE": 1, "ANCIENT": 2}
    df_novel = df_classified[
        df_classified["numt_class"] == "NOVEL_NUMT_CANDIDATE"
    ].copy()
    df_novel["_age_rank"] = df_novel["insertion_age_estimate"].map(_age_order).fillna(3)
    df_novel = df_novel.sort_values(
        ["_age_rank", "mito_combined_score"], ascending=[True, False], na_position="last"
    ).drop(columns=["_age_rank"])

    df_confirmed = df_classified[
        df_classified["numt_class"].isin(["CONFIRMED_NUMT", "KNOWN_REFERENCE_NUMT"])
    ].sort_values("mito_combined_score", ascending=False, na_position="last")

    df_mito_ref = df_classified[
        df_classified["numt_class"] == "DIRECT_MITO_REF"
    ]

    sheets["NOVEL_CANDIDATES"]  = df_novel
    sheets["MITO_AND_NUCLEAR"]  = df_confirmed
    sheets["DIRECT_MITO_REF"]   = df_mito_ref

    log(f"\n  NOVEL_NUMT_CANDIDATE  : {len(df_novel):,} variants  ← cibles principales")
    log(f"  MITO_AND_NUCLEAR      : {len(df_confirmed):,} variants")
    log(f"  DIRECT_MITO_REF       : {len(df_mito_ref):,} variants  (annotations HGVS explicites)")

    # Gènes de maladies dans les candidats
    if not df_novel.empty and "is_disease_gene" in df_novel.columns:
        disease_candidates = df_novel[df_novel["is_disease_gene"] == True]
        if not disease_candidates.empty:
            log(f"\n  *** ATTENTION : {len(disease_candidates)} candidat(s) dans des gènes de maladies connues ***")
            for _, row in disease_candidates.iterrows():
                gene = row.get("GeneSymbol", "?")
                aid  = row.get("AlleleID", "?")
                pident = row.get("mito_pident", "?")
                log(f"      AlleleID {aid}  gene={gene}  pident_mito={pident}%")

    # Variants somatiques parmi les candidats NOVEL
    if not df_novel.empty and "origin_group" in df_novel.columns:
        somatic_novel = df_novel[df_novel["origin_group"] == "somatic"]
        if not somatic_novel.empty:
            log(f"\n  *** ATTENTION : {len(somatic_novel)} NOVEL candidat(s) d'ORIGINE SOMATIQUE ***")
            log("      Interprétation différente des candidats germinaux :")
            log("      → Pas de calcul de risque de récurrence")
            log("      → Orienter vers COSMIC / TCGA pour contexte somatique")
            log("      → Référence : Singh et al. 2017 (numtogenèse dans le cancer)")
            for _, row in somatic_novel.iterrows():
                gene  = row.get("GeneSymbol", "?")
                aid   = row.get("AlleleID", "?")
                orig  = row.get("Origin", "?")
                log(f"      AlleleID {aid}  gene={gene}  Origin={orig}")
        else:
            log(f"\n  Origine : tous les candidats NOVEL sont germinaux "
                f"({df_novel['origin_group'].value_counts().to_dict()})")

    # ── ÉTAPE 7 : analyse du mécanisme d'insertion ──────────────────────
    log("")
    log("─── Étape 7 : Mécanisme d'insertion ───")
    df_mechanism = run_mechanism_step(df_classified, df_mito, args.email)
    sheets["7_Mechanism"] = df_mechanism

    if not df_mechanism.empty:
        # Résumé des mécanismes prédits
        mech_counts = df_mechanism["predicted_mechanism"].value_counts()
        for mech, count in mech_counts.items():
            log(f"  {mech:<20} {count:>4,}")

        # Feuille statistiques (tuple → traitement spécial dans write_excel)
        sheets["8_Statistics"] = (df_mechanism, df_novel, df_confirmed)

        # Ajouter les colonnes mécanisme à NOVEL_CANDIDATES pour vue consolidée
        mech_cols = [
            "AlleleID", "predicted_mechanism", "mechanism_confidence",
            "insertion_orientation", "microhomology_found", "microhomology_length",
            "tsd_found", "l1_site_found", "l1_within_5kb",
            "at_content_combined", "dna_curvature_mean",
            "nearest_repeat_class", "nearest_repeat_dist_bp",
            "recombination_rate_cM_Mb",
        ]
        # Keep only columns that actually exist (l1_within_5kb etc. may be absent
        # if the pipeline ran on an older version of 06_mechanism.py)
        mech_cols = [c for c in mech_cols if c in df_mechanism.columns]
        mech_for_merge = df_mechanism[mech_cols].copy()
        mech_for_merge["AlleleID"] = mech_for_merge["AlleleID"].astype(str)

        if not df_novel.empty:
            df_novel_enriched = df_novel.copy()
            df_novel_enriched["AlleleID"] = df_novel_enriched["AlleleID"].astype(str)
            df_novel_enriched = df_novel_enriched.merge(
                mech_for_merge, on="AlleleID", how="left"
            )
            sheets["NOVEL_CANDIDATES"] = df_novel_enriched

    # ── Écriture Excel ───────────────────────────────────────────────────
    log("")
    write_excel(sheets, args.output)

    log("")
    log("=" * 65)
    log("Pipeline terminée.")
    log(f"Résultats : {args.output}")
    log("=" * 65)


if __name__ == "__main__":
    main()
